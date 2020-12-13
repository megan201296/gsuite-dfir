from __future__ import print_function
import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import pandas as pd
import geoip2.database
import sys
from openpyxl import load_workbook
import time
import json
import argparse

class Google(object):
    """
    Class for doing bulk of operations related to Google Workspace DFIR activities
    """

    def __init__(self):
        self.output = args.output
        self.geolocate_db = config['geolocate_db']
        self.token_path = config['token_path']
        self.creds_path = config['creds_path']
        self.service = self.google_session()

    def google_session(self):
        """
        Establish connection to Google Wrospace.
        """
        creds = None
        SCOPES = ['https://www.googleapis.com/auth/admin.reports.audit.readonly']

        # The file token.pickle stores the user's access and refresh tokens, and is
        # created automatically when the authorization flow completes for the first
        # time.
        if os.path.exists(self.token_path):
            with open(self.token_path, 'rb') as token:
                creds = pickle.load(token)
        # If there are no (valid) credentials available, let the user log in.
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    self.creds_path, SCOPES)
                creds = flow.run_local_server(port=0)
            # Save the credentials for the next run
            with open(self.token_path, 'wb') as token:
                pickle.dump(creds, token)

        service = build('admin', 'reports_v1', credentials=creds)

        return service

    def get_login_activity(self):

        # Call the Admin SDK Reports API
        results = self.service.activities().list(
            userKey='all', applicationName='login').execute()
        activities = results.get('items', [])

        # save logs to Pandas data frame and clean up data
        df_activities = pd.json_normalize(activities)
        df_events = pd.json_normalize(data=activities, record_path=['events'])
        df_logs = df_activities.join(df_events)
        df_params = df_logs['parameters'].apply(pd.Series)
        df_params =  df_params.rename(columns = lambda x : 'param_' + str(x))
        df_logs = df_activities.join(df_params)
        df_logs = df_logs.drop(columns=['events', 'kind', 'etag', 'id.uniqueQualifier', 'id.customerId', 'actor.profileId'])
        df_logs = df_logs.rename(columns={"id.time": "timestamp", "actor.email": "userEmail", "id.applicationName": "applicationName"})
        df_logs['loginCountry'] = df_logs['ipAddress'].map(lambda ipAddress: self.get_geoip(ipAddress)[0])
        df_logs['loginCity'] = df_logs['ipAddress'].map(lambda ipAddress: self.get_geoip(ipAddress)[1])
        df_logs.to_excel(self.output, "Login Activity", index=False)

    def get_drive_activity(self):

        # Call the Admin SDK Reports API
        results = self.service.activities().list(
            userKey='all', applicationName='drive').execute()
        activities = results.get('items', [])

        # save logs to Pandas data frame and clean up data
        df_activities = pd.json_normalize(activities)
        df_events = pd.json_normalize(data=activities, record_path=['events'])
        df_logs = df_activities.join(df_events)
        df_params = df_logs['parameters'].apply(pd.Series)
        df_params =  df_params.rename(columns = lambda x : 'param_' + str(x))
        df_logs = df_activities.join(df_params)
        df_logs = df_logs.drop(columns=['events', 'kind', 'etag', 'id.uniqueQualifier', 'actor.profileId', 'id.customerId'])
        df_logs = df_logs.rename(columns={"id.time": "timestamp", "actor.email": "userEmail", "id.applicationName": "applicationName"})
        df_logs['loginCountry'] = df_logs['ipAddress'].apply(lambda ipAddress: self.get_geoip(ipAddress)[0] if type(ipAddress) == str  else "")
        df_logs['loginCity'] = df_logs['ipAddress'].apply(lambda ipAddress: self.get_geoip(ipAddress)[1] if type(ipAddress) == str  else "")

        with pd.ExcelWriter(self.output, engine='openpyxl', mode='a') as writer:
            writer.book = load_workbook(self.output)
            df_logs.to_excel(writer, "Google Drive Activity", index=False)

    def get_admin_activity(self):

        # Call the Admin SDK Reports API
        results = self.service.activities().list(
            userKey='all', applicationName='admin').execute()
        activities = results.get('items', [])

        # save logs to Pandas data frame and clean up data
        df_activities = pd.json_normalize(activities)
        df_events = pd.json_normalize(data=activities, record_path=['events'])
        df_logs = df_activities.join(df_events)
        df_params = df_logs['parameters'].apply(pd.Series)
        df_params =  df_params.rename(columns = lambda x : 'param_' + str(x))
        df_logs = df_activities.join(df_params)
        df_logs = df_logs.drop(columns=['events', 'kind', 'etag', 'id.uniqueQualifier', 'actor.profileId', 'id.customerId'])
        df_logs = df_logs.rename(columns={"id.time": "timestamp", "actor.email": "userEmail", "id.applicationName": "applicationName"})
        df_logs['loginCountry'] = df_logs['ipAddress'].apply(lambda ipAddress: self.get_geoip(ipAddress)[0] if type(ipAddress) == str  else "")
        df_logs['loginCity'] = df_logs['ipAddress'].apply(lambda ipAddress: self.get_geoip(ipAddress)[1] if type(ipAddress) == str  else "")        

        with pd.ExcelWriter(self.output, engine='openpyxl', mode='a') as writer:
            writer.book = load_workbook(self.output)
            df_logs.to_excel(writer, "Admin Activity", index=False)

    def get_user_activity(self):

        # Call the Admin SDK Reports API
        results = self.service.activities().list(
            userKey='all', applicationName='user_accounts').execute()
        activities = results.get('items', [])
        # save logs to Pandas data frame and clean up data
        df_activities = pd.json_normalize(activities)
        df_events = pd.json_normalize(data=activities, record_path=['events'])
        df_logs = df_activities.join(df_events)
        df_logs = df_logs.drop(columns=['events', 'kind', 'etag', 'id.uniqueQualifier', 'actor.profileId', 'id.customerId'])
        df_logs = df_logs.rename(columns={"id.time": "timestamp", "actor.email": "userEmail", "id.applicationName": "applicationName"})
        df_logs['loginCountry'] = df_logs['ipAddress'].apply(lambda ipAddress: self.get_geoip(ipAddress)[0] if type(ipAddress) == str  else "")
        df_logs['loginCity'] = df_logs['ipAddress'].apply(lambda ipAddress: self.get_geoip(ipAddress)[1] if type(ipAddress) == str  else "")
        df_logs = df_logs[["timestamp", "userEmail", "ipAddress", "loginCountry", "loginCity", "applicationName", "actor.callerType", "type", "name"]]

        with pd.ExcelWriter(self.output, engine='openpyxl', mode='a') as writer:
            writer.book = load_workbook(self.output)
            df_logs.to_excel(writer, "User Activity", index=False)

    def get_geoip(self, ipAddress):
        reader = geoip2.database.Reader(
            self.geolocate_db)
        response = reader.city(ipAddress)
        return [response.country.iso_code, response.city.name]
    

    def timeline(self):
        sheet_to_df_map = pd.read_excel(self.output, sheet_name=None)
        timeline = pd.concat(sheet_to_df_map, axis=0, ignore_index=True)
        
        with pd.ExcelWriter(self.output, engine='openpyxl', mode='a') as writer:
            writer.book = load_workbook(self.output)
            timeline.to_excel(writer, "All", index=False)

def parse_args():
    parser = argparse.ArgumentParser(
        description="Retrieve and process Google Workspace logs"
    )

    parser.add_argument(
        "-o", "--output",
        help="Output path"
    )

    args = parser.parse_args()

    return args

start_time = time.time()

args = parse_args()

with open("config.json") as json_data_file:
    config = json.load(json_data_file)

google = Google()

vars(google)

google.get_login_activity()
google.get_drive_activity()
google.get_admin_activity()
google.get_user_activity()
google.timeline()

elapsed = time.time() - start_time
print(f'Total execution time: {elapsed}')